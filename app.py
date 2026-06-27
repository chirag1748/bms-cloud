"""
app.py — BMS Cloud
Flask application. All routes in one file for simplicity.
"""

import os
import io
import secrets
from datetime import datetime, timedelta
from functools import wraps

from flask import (Flask, request, jsonify, render_template,
                   send_from_directory, send_file)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dotenv import load_dotenv

load_dotenv()
import database as db

app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))

db.init_db()

# ── MongoDB-backed sessions (survives server restarts) ────────
SESSION_H = 12


def _make_token(username, role):
    token   = secrets.token_hex(32)
    expires = datetime.utcnow() + timedelta(hours=SESSION_H)
    mdb = db.get_db()
    mdb.sessions.create_index("expires", expireAfterSeconds=0)
    mdb.sessions.insert_one({
        "token":    token,
        "username": username,
        "role":     role,
        "expires":  expires
    })
    return token


def _get_sess(token):
    if not token:
        return None
    mdb  = db.get_db()
    sess = mdb.sessions.find_one({
        "token":   token,
        "expires": {"$gt": datetime.utcnow()}
    })
    if sess:
        return {"username": sess["username"], "role": sess["role"]}
    return None


def _delete_token(token):
    if token:
        db.get_db().sessions.delete_one({"token": token})


def _token_from_req():
    return (request.headers.get("X-Session-Token")
            or request.cookies.get("bms_token")
            or (request.get_json(silent=True) or {}).get("_token"))


def require_auth(role=None):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            sess = _get_sess(_token_from_req())
            if not sess:
                return jsonify({"error": "Unauthorized"}), 401
            if role and sess["role"] != role and sess["role"] != "admin":
                return jsonify({"error": "Forbidden"}), 403
            request.bms_sess = sess
            return f(*args, **kwargs)
        return wrapper
    return decorator


# ── PWA static ────────────────────────────────────────────────
@app.route("/manifest.json")
def manifest():
    return jsonify({
        "name":             "BMS – Breakdown Management",
        "short_name":       "BMS",
        "start_url":        "/",
        "display":          "standalone",
        "background_color": "#1B3A6B",
        "theme_color":      "#1B3A6B",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    })


@app.route("/sw.js")
def sw():
    return send_from_directory("static", "sw.js",
                               mimetype="application/javascript")


# ── Auth routes ───────────────────────────────────────────────
@app.route("/api/login", methods=["POST"])
def api_login():
    d        = request.get_json() or {}
    username = (d.get("username") or "").strip()
    password = d.get("password") or ""
    user     = db.authenticate_user(username, password)
    if not user:
        return jsonify({"error": "Invalid username or password"}), 401
    token = _make_token(user["username"], user["role"])
    resp  = jsonify({"token": token, "username": user["username"],
                     "role": user["role"]})
    resp.set_cookie("bms_token", token, httponly=True,
                    samesite="Lax", max_age=SESSION_H * 3600)
    return resp, 200


@app.route("/api/logout", methods=["POST"])
def api_logout():
    token = _token_from_req()
    _delete_token(token)
    resp = jsonify({"status": "ok"})
    resp.delete_cookie("bms_token")
    return resp, 200


@app.route("/api/me", methods=["GET"])
@require_auth()
def api_me():
    return jsonify(request.bms_sess), 200


# ── Master data ───────────────────────────────────────────────
@app.route("/api/master_data", methods=["GET"])
@require_auth()
def api_master_data():
    return jsonify({
        "vehicles":     db.get_vehicles(),
        "technicians":  db.get_technicians(),
        "electricians": db.get_electricians(),
        "drivers":      db.get_drivers(),
    }), 200


@app.route("/api/admin/master/<mtype>", methods=["GET"])
@require_auth("admin")
def admin_get_master(mtype):
    if mtype not in ("vehicle", "technician", "electrician", "driver"):
        return jsonify({"error": "Invalid type"}), 400
    return jsonify(db.get_all_master(mtype)), 200


@app.route("/api/admin/master/<mtype>", methods=["POST"])
@require_auth("admin")
def admin_add_master(mtype):
    if mtype not in ("vehicle", "technician", "electrician", "driver"):
        return jsonify({"error": "Invalid type"}), 400
    name = (request.get_json() or {}).get("name", "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    db.add_master(mtype, name)
    return jsonify({"status": "ok"}), 200


@app.route("/api/admin/master/<item_id>/toggle", methods=["POST"])
@require_auth("admin")
def admin_toggle_master(item_id):
    active = (request.get_json() or {}).get("active", True)
    db.toggle_master(item_id, active)
    return jsonify({"status": "ok"}), 200


@app.route("/api/admin/master/<item_id>", methods=["DELETE"])
@require_auth("admin")
def admin_delete_master(item_id):
    db.delete_master(item_id)
    return jsonify({"status": "ok"}), 200


# ── Users (admin only) ────────────────────────────────────────
@app.route("/api/admin/users", methods=["GET"])
@require_auth("admin")
def admin_users():
    users = db.get_all_users()
    for u in users:
        u["id"] = str(u.pop("_id"))
        u.pop("password_hash", None)
    return jsonify(users), 200


@app.route("/api/admin/users", methods=["POST"])
@require_auth("admin")
def admin_add_user():
    d    = request.get_json() or {}
    uname = (d.get("username") or "").strip()
    pw    = (d.get("password") or "").strip()
    role  = d.get("role", "bd_incharge")
    if not uname or not pw:
        return jsonify({"error": "Username and password required"}), 400
    if role not in ("admin", "bd_incharge"):
        return jsonify({"error": "Invalid role"}), 400
    try:
        db.add_user(uname, pw, role)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    return jsonify({"status": "ok"}), 200


@app.route("/api/admin/users/<user_id>", methods=["DELETE"])
@require_auth("admin")
def admin_delete_user(user_id):
    db.delete_user(user_id)
    return jsonify({"status": "ok"}), 200


@app.route("/api/admin/users/<user_id>/password", methods=["POST"])
@require_auth("admin")
def admin_change_pw(user_id):
    pw = (request.get_json() or {}).get("password", "").strip()
    if not pw:
        return jsonify({"error": "Password required"}), 400
    db.change_password(user_id, pw)
    return jsonify({"status": "ok"}), 200


# ── Excel export helpers ────────────────────────────────────────
def _autosize_and_header(ws, headers):
    header_fill = PatternFill("solid", start_color="1B3A6B", end_color="1B3A6B")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _autosize_and_header_at_row(ws, headers, row):
    header_fill = PatternFill("solid", start_color="1B3A6B", end_color="1B3A6B")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def _autowidth(ws, headers):
    for col_idx, title in enumerate(headers, start=1):
        max_len = len(title)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)


def _send_workbook(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/tickets/export", methods=["GET"])
@require_auth()
def api_tickets_export():
    token_no   = request.args.get("token_no", "")
    rc_no      = request.args.get("rc_no", "")
    technician = request.args.get("technician", "")
    from_date  = request.args.get("from_date")
    to_date    = request.args.get("to_date")
    rows = db.get_tickets_for_export(token_no=token_no, rc_no=rc_no,
                                     technician=technician,
                                     from_date=from_date, to_date=to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    headers = ["#", "Token", "RC No", "Technician", "Driver", "Vehicle",
              "Electrician", "Assigned", "Reach", "Resolution",
              "Reach Delay Reason", "Resolution Delay Reason", "Status", "Submitted By"]
    _autosize_and_header(ws, headers)
    for i, t in enumerate(rows, start=1):
        ws.append([
            i, t.get("token_no") or "-", t.get("rc_no") or "-",
            t.get("technician") or "-", t.get("driver") or "-",
            t.get("vehicle_used") or "-", t.get("electrician") or "-",
            db.fmt_display_dt(t.get("assigned_time")) or "-",
            db.fmt_display_dt(t.get("reach_time")) or "-",
            db.fmt_display_dt(t.get("resolution_time")) or "-",
            t.get("reach_delay_reason") or "", t.get("resolution_delay_reason") or "",
            t.get("status") or "-", t.get("submitted_by") or "-",
        ])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri")
    _autowidth(ws, headers)

    fname = f"bms_tickets_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return _send_workbook(wb, fname)


@app.route("/api/reports/export", methods=["GET"])
@require_auth("admin")
def api_reports_export():
    from_date = request.args.get("from_date",
                                 datetime.utcnow().replace(day=1).strftime("%Y-%m-%d"))
    to_date   = request.args.get("to_date",
                                 datetime.utcnow().strftime("%Y-%m-%d"))
    summary = db.get_summary_counts(from_date, to_date)
    tech_rows = db.get_technician_performance(from_date, to_date)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["BMS Dashboard Report"])
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color="1B3A6B")
    ws1.append([f"Period: {from_date} to {to_date}"])
    ws1.append([])
    headers1 = ["Metric", "Value"]
    _autosize_and_header_at_row(ws1, headers1, row=4)
    ws1.append(["Total Tickets", summary["total"]])
    ws1.append(["CONF", summary["conf"]])
    ws1.append(["CONF %", f"{summary['conf_pct']}%"])
    ws1.append(["NC", summary["nc"]])
    ws1.append(["NC %", f"{summary['nc_pct']}%"])
    for row in ws1.iter_rows(min_row=5):
        for cell in row:
            cell.font = Font(name="Calibri")
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 16

    ws2 = wb.create_sheet("Technician Performance")
    headers2 = ["Technician", "Total", "CONF", "NC", "CONF %"]
    _autosize_and_header(ws2, headers2)
    for r in tech_rows:
        ws2.append([r["technician"], r["total"], r["conf"], r["nc"], f"{r['conf_pct']}%"])
    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri")
    _autowidth(ws2, headers2)

    fname = f"bms_report_{from_date}_to_{to_date}.xlsx"
    return _send_workbook(wb, fname)


# ── Tickets ───────────────────────────────────────────────────
@app.route("/api/tickets", methods=["GET"])
@require_auth()
def api_tickets():
    token_no   = request.args.get("token_no", "")
    rc_no      = request.args.get("rc_no", "")
    technician = request.args.get("technician", "")
    from_date  = request.args.get("from_date")
    to_date    = request.args.get("to_date")
    rows = db.search_tickets(token_no=token_no, rc_no=rc_no,
                              technician=technician,
                              from_date=from_date, to_date=to_date)
    return jsonify(rows), 200


@app.route("/api/tickets", methods=["POST"])
@require_auth()
def api_add_ticket():
    data = request.get_json() or {}
    data["submitted_by"] = request.bms_sess["username"]
    try:
        result, ticket_id = db.upload_ticket_safe(data)
        return jsonify({"status": result, "ticket_id": ticket_id}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/tickets/<ticket_id>", methods=["PUT"])
@require_auth("admin")
def api_update_ticket(ticket_id):
    data = request.get_json() or {}
    db.update_ticket(ticket_id, data)
    return jsonify({"status": "ok"}), 200


@app.route("/api/tickets/<ticket_id>", methods=["DELETE"])
@require_auth("admin")
def api_delete_ticket(ticket_id):
    db.delete_ticket(ticket_id)
    return jsonify({"status": "ok"}), 200


# ── Reports (admin only) ──────────────────────────────────────
@app.route("/api/reports/summary", methods=["GET"])
@require_auth("admin")
def api_summary():
    from_date = request.args.get("from_date",
                                 datetime.utcnow().replace(day=1).strftime("%Y-%m-%d"))
    to_date   = request.args.get("to_date",
                                 datetime.utcnow().strftime("%Y-%m-%d"))
    return jsonify({
        "summary":     db.get_summary_counts(from_date, to_date),
        "technicians": db.get_technician_performance(from_date, to_date),
    }), 200


# ── App shell (catch-all → SPA) ───────────────────────────────
@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def catch_all(path):
    return render_template("app.html")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5055))
    app.run(host="0.0.0.0", port=port, debug=False)
